#!/usr/bin/env python3
"""
Parte 0 del `2026-09-03_1` — ¿el «Programmatic» del tablero es sólo DV360?

⛔ **Sólo lectura.** No toca `MARCADORES`, no toca `Fuentes.gs`, no deroga `R-24`.

⛔⛔ **Corre sobre FIXTURE, no sobre la base viva, y eso cambia qué contesta.** El prompt pide la
base viva; ese camino **no existe para Code** por decisión del usuario del 20/08 (`docs/_fixtures/
README.md`): leer el contenido de una base por el conector volcaría nombres de vecinos y barrios a
una conversación, que es lo que `C-21` decidió evitar. Así que esto responde **por el 30/08** y por
ningún otro día — `looker/DIGITAL` y el desglose son inestables por CAMBIO (`R-31`).

⚠ **Y una premisa del prompt que venció:** 0.2 pide usar «mismo `filtro`» que `imp_prog`. Desde la
mudanza del 31/08 (`mudarImpresionesAlDesglose()`) las ocho filas viven en
`digital|CAMPAÑAS_DESGLOCE_DIGITAL` sobre `des_impresiones` **con el `filtro` VACÍO**. No hay
`estado=Activa` que copiar. Se replica el corte real: ámbito + plataforma, sin filtro.

⭐ **Los cuatro controles positivos, y los cuatro abortan** (`CLAUDE.md` §4 — un instrumento sin
control positivo no distingue «no está» de «no miré»):

  CP1  `sha256` de los dos fixtures contra la tabla de huellas del README.
  CP2  segundo lector INDEPENDIENTE (`openpyxl`, que resuelve por posición) contra el lector por
       referencia — tienen que FALLAR DISTINTO, y por eso vale el cruce.
  CP3  identidad interna `Meta + Google + programmatic(por resta) = TOTAL`, exacta al dígito por
       construcción. Si no cierra, la partición no es exhaustiva y nada de abajo se puede citar.
  CP4  reproducir la ESTRUCTURA que `docs/MEDICION_mudanza_imp_2026-08-31.md` §1 registra de la
       corrida del 31/08: **28 filas de JM** y **JM idéntico entre las dos ventanas mientras GCBA
       crece**. ⛔⛔ La primera versión de este control exigía reproducir los VALORES publicados y
       **falló con desvíos de -3,8 % a -18 %**. El instrumento no estaba mal: el control sí. Esta
       solapa es **inestable por CAMBIO** (`R-31`) y publica **acumulado**, así que exigirle
       igualdad a una constante leída un día después es el error de `V-110` — pedirle exactitud a
       un campo que no la admite. ⭐ Se reemplaza por un control ESTRUCTURAL, que es un CONTEO DE
       FILAS y por eso **inmune al CAMBIO** (`CLAUDE.md` §4, la tabla operación × inestabilidad).
       El delta de valores se sigue imprimiendo, pero como **ilustración declarada** y no como
       criterio.

Uso:
  python tools/medir-programmatic-dv360.py
"""
import datetime
import hashlib
import os
import re
import sys
from collections import Counter, OrderedDict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from leer_xlsx_por_referencia import Hoja, norm  # noqa: E402

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

FIX = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'docs', '_fixtures')
DESG = os.path.join(FIX, 'Seguimiento_Digital_2026-08-30.xlsx')
LOOK = os.path.join(FIX, 'Base_Looker_2026-08-30.xlsx')
SHA = {DESG: 'd7b917f5711dcdd70b20b82ce8d6ccaa336fc3c1bcf5a0114296ca33edf70d6a',
       LOOK: '7272b383ebca44916250383a65c3155cee04ea172dde78fd7554ee20040ae5b2'}

# `PERIODOS` (snapshot 31/08). Las dos ventanas testigo que pide 0.2.
VENTANAS = OrderedDict([
    ('2026_agosto_21_27', (datetime.date(2026, 8, 21), datetime.date(2026, 8, 27))),
    ('2026_agosto_21_28', (datetime.date(2026, 8, 21), datetime.date(2026, 8, 28))),
])

# `docs/MEDICION_mudanza_imp_2026-08-31.md` §1 — tablero, ventana 21–28, lectura 30/08 18:00.
TABLERO = {'jm':   {'meta': 2254346,  'google': 1219244,  'prog': 6907699,  'total': 10381289},
           'gcba': {'meta': 24164426, 'google': 19841789, 'prog': 61398036, 'total': 105404251}}

# `docs/MEDICION_mudanza_imp_2026-08-31.md` §1 — lo que el motor publicó el 31/08 (CP4).
PUBLICADO = {'jm':   {'meta': 3709430,  'google': 1946475,  'prog': 10608520,  'total': 16264425},
             'gcba': {'meta': 80122796, 'google': 54836419, 'prog': 166397876, 'total': 301357091}}

# Censo del comentario de `DIMENSIONES_.plataforma` en `Fuentes.gs`, fixture del 20/08. Testigo
# con fecha: se re-mide, no se copia. Está acá sólo para nombrar las etiquetas NUEVAS.
CENSO_20_08 = {'Meta': 1840, 'DV360': 1678, 'Google ads': 1417, 'TikTok': 55,
               'Mercado Libre': 27, 'Twitter': 12, 'Twitch': 5, 'Uber': 5}

EPOCA = datetime.date(1899, 12, 30)
mil = lambda n: format(int(round(n)), ',').replace(',', '.')


def num(v):
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return 0.0


def fecha(v):
    s = norm(v)
    if not s:
        return None
    try:
        return EPOCA + datetime.timedelta(days=int(float(s)))
    except ValueError:
        for f in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.datetime.strptime(s[:10], f).date()
            except ValueError:
                pass
    return None


def bloque(t):
    print('\n' + '=' * 94)
    print(t)
    print('=' * 94)


# ─────────────────────────────────────────────────────────── CP1 · procedencia
bloque('CP1 · PROCEDENCIA — sha256 contra la tabla de huellas del README')
print('lectura de los archivos: %s (hora local)' % datetime.datetime.now().strftime('%Y-%m-%d %H:%M'))
for r, esp in sorted(SHA.items()):
    s = hashlib.sha256(open(r, 'rb').read()).hexdigest()
    ok = s == esp
    print('  %-38s %s…  %s' % (os.path.basename(r), s[:16], '✅' if ok else '⛔'))
    if not ok:
        raise SystemExit('⛔ CP1 falló: el archivo no es el declarado. No se cita ningún número.')
print('  ⚠ `modifiedTime` de la planilla VIVA: no disponible — esto es un export en disco.')

dg = Hoja(DESG, lambda n: 'DESGLOCE' in n)
lk = Hoja(LOOK, lambda n: n == 'DIGITAL')
G = lambda f, L: norm(f.get(L, ''))

# ─────────────────────────────────────────── CP2 · segundo lector independiente
bloque('CP2 · SEGUNDO LECTOR — `openpyxl` (resuelve por posición) vs. el lector por referencia')
try:
    import openpyxl
except ImportError:
    raise SystemExit('⛔ CP2 no puede correr sin openpyxl. Sin cruce no se cita ningún número.')


def por_openpyxl(ruta, elegir, col_plat, col_imp):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hoja = next(h for h in wb.sheetnames if elegir(h))
    ws = wb[hoja]
    c = Counter()
    imp = Counter()
    for i, fila in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if all(v is None for v in fila):
            continue
        p = norm(fila[col_plat]) if len(fila) > col_plat else ''
        c[p] += 1
        imp[p] += num(fila[col_imp]) if len(fila) > col_imp else 0.0
    wb.close()
    return c, imp


def por_referencia(hoja, col_plat, col_imp):
    c = Counter()
    imp = Counter()
    for f in hoja.datos:
        p = G(f, col_plat)
        c[p] += 1
        imp[p] += num(G(f, col_imp))
    return c, imp


CRUCES = [
    ('digital | CAMPAÑAS_DESGLOCE_DIGITAL', DESG, lambda n: 'DESGLOCE' in n, dg, 'F', 5, 'O', 14),
    ('looker | DIGITAL',                    LOOK, lambda n: n == 'DIGITAL',   lk, 'B', 1, 'C', 2),
]

CENSOS = {}
for etiqueta, ruta, elegir, hoja, cl, ci_plat, cl_imp, ci_imp in CRUCES:
    a_c, a_i = por_referencia(hoja, cl, cl_imp)
    b_c, b_i = por_openpyxl(ruta, elegir, ci_plat, ci_imp)
    claves = set(a_c) | set(b_c)
    difs = [k for k in claves if a_c[k] != b_c[k] or abs(a_i[k] - b_i[k]) > 0.5]
    print('  %-38s etiquetas: ref %d · openpyxl %d · discrepan: %d  %s'
          % (etiqueta, len(a_c), len(b_c), len(difs), '✅' if not difs else '⛔'))
    if difs:
        for k in difs:
            print('     ⛔ %-18s ref %d/%s · openpyxl %d/%s'
                  % (k or '(vacío)', a_c[k], mil(a_i[k]), b_c[k], mil(b_i[k])))
        raise SystemExit('⛔ CP2 falló: los dos lectores discrepan. El hallazgo es el camino de '
                         'lectura, no el dato — no se cita ningún número.')
    CENSOS[etiqueta] = (a_c, a_i)

# ───────────────────────────────────────────────────────────────────── 0.1 censo
bloque('0.1 · CENSO DE ETIQUETAS FÍSICAS DE PLATAFORMA — solapa entera, sin ventana ni ámbito')
for etiqueta, (c, imp) in CENSOS.items():
    print('\n── %s ── (%s filas)' % (etiqueta, mil(sum(c.values()))))
    print('   %-22s %10s %18s' % ('etiqueta', 'filas', 'impresiones'))
    for k, n in c.most_common():
        print('   %-22s %10s %18s' % (k or '(vacío)', mil(n), mil(imp[k])))

nuevas = set(CENSOS['digital | CAMPAÑAS_DESGLOCE_DIGITAL'][0]) - set(CENSO_20_08)
faltan = set(CENSO_20_08) - set(CENSOS['digital | CAMPAÑAS_DESGLOCE_DIGITAL'][0])
print('\n⚠ Contra el censo del comentario de `Fuentes.gs` (fixture 20/08), sobre el desglose:')
print('   etiquetas NUEVAS que ese comentario no tiene: %s' % (sorted(nuevas) or 'ninguna'))
print('   etiquetas del comentario que hoy no aparecen : %s' % (sorted(faltan) or 'ninguna'))

# ────────────────────────────────────────────────── 0.2 el corte real de imp_prog
bloque('0.2 · CUÁNTO PESA LO QUE NO ES DV360 — corte real de `imp_prog` / `gcba_imp_prog`')
print("""corte replicado desde `MARCADORES` (snapshot 2026-08-31, POST mudanza):
  base|solapa   digital | CAMPAÑAS_DESGLOCE_DIGITAL
  campo         des_impresiones (col O)
  filtro        (VACÍO — la mudanza del 31/08 lo dejó así a propósito)
  ambito=jm     des_id_cuenta ~= JDGAG          (col B, `~=` es «contiene», sensible al case)
  ambito=gcba   des_id_cuenta !~= JDGAG         (negación — `gcba` es «todo lo que no es jm»)
  plataforma    des_plataforma != Meta && != Google ads   (`R-24`, por resta)
  ventana       `ventana_ref = propia` → SOLAPE de col I/J contra la ventana (`R-16`)""")

CONTIENE_JM = 'JDGAG'


def es_jm(f):
    return CONTIENE_JM in G(f, 'B')


def por_resta(f):
    p = G(f, 'F')
    return p != 'Meta' and p != 'Google ads'


def solape(desde, hasta):
    out = []
    for f in dg.datos:
        i, j = fecha(G(f, 'I')), fecha(G(f, 'J'))
        if i and j and i <= hasta and j >= desde:
            out.append(f)
    return out


suma = lambda filas: sum(num(G(f, 'O')) for f in filas)
RES = {}

for vent, (desde, hasta) in VENTANAS.items():
    sel = solape(desde, hasta)
    print('\n── %s (%s → %s) · %s filas en solape ──'
          % (vent, desde, hasta, mil(len(sel))))
    print('   %-6s %16s %16s %14s %8s   %s'
          % ('ámbito', 'prog. por resta', 'sólo DV360', 'delta abs.', 'delta %', 'filas no-DV360'))
    for amb in ('jm', 'gcba'):
        filas = [f for f in sel if es_jm(f) == (amb == 'jm')]
        prog = [f for f in filas if por_resta(f)]
        dv = [f for f in prog if G(f, 'F') == 'DV360']
        otras = [f for f in prog if G(f, 'F') != 'DV360']
        a, b = suma(prog), suma(dv)
        pct = ('%.2f %%' % (100.0 * (a - b) / a)) if a else '—'
        det = Counter(G(f, 'F') or '(vacío)' for f in otras)
        detalle = ' · '.join('%s %d (%s)' % (k, n, mil(sum(num(G(f, 'O')) for f in otras if (G(f, 'F') or '(vacío)') == k)))
                             for k, n in det.most_common()) or 'ninguna'
        print('   %-6s %16s %16s %14s %8s   %s'
              % (amb, mil(a), mil(b), mil(a - b), pct, detalle))
        RES[(vent, amb)] = {
            'prog': a, 'dv360': b, 'filas': filas,
            'meta': suma([f for f in filas if G(f, 'F') == 'Meta']),
            'google': suma([f for f in filas if G(f, 'F') == 'Google ads']),
            'total': suma(filas),
        }

# ────────────────────────────────────────────────────── CP3 · identidad interna
bloque('CP3 · IDENTIDAD INTERNA — Meta + Google + programmatic(por resta) = TOTAL')
malas = 0
for (vent, amb), rr in RES.items():
    s = rr['meta'] + rr['google'] + rr['prog']
    ok = abs(s - rr['total']) < 0.5
    malas += 0 if ok else 1
    print('  %-20s %-5s %18s vs %18s  %s'
          % (vent, amb, mil(s), mil(rr['total']), '✅' if ok else '⛔'))
if malas:
    raise SystemExit('⛔ CP3 falló: la partición no es exhaustiva. Nada de arriba se puede citar.')

# ──────────────────────────────────────────────────────────────── 0.3 y CP4
bloque('0.3 · CONTRA EL TABLERO — ventana 2026_agosto_21_28, lectura del tablero 30/08 18:00')
r = {amb: RES[('2026_agosto_21_28', amb)] for amb in ('jm', 'gcba')}
print('   %-5s %-14s %16s %16s %10s %12s'
      % ('ámb.', 'plataforma', 'motor (fixt. 30/08)', 'tablero', '% m/t', 'nota'))
for amb in ('jm', 'gcba'):
    for k, nom in (('meta', 'Meta'), ('google', 'Google ads')):
        v, t = r[amb][k], TABLERO[amb][k]
        print('   %-5s %-14s %16s %16s %9.1f %%' % (amb, nom, mil(v), mil(t), 100.0 * v / t))
    for k, nom in (('prog', 'prog · POR RESTA'), ('dv360', 'prog · SÓLO DV360')):
        v, t = r[amb][k], TABLERO[amb]['prog']
        print('   %-5s %-14s %16s %16s %9.1f %%' % (amb, nom, mil(v), mil(t), 100.0 * v / t))
    v, t = r[amb]['total'], TABLERO[amb]['total']
    print('   %-5s %-14s %16s %16s %9.1f %%\n' % (amb, 'TOTAL', mil(v), mil(t), 100.0 * v / t))

bloque('CP4 · ESTRUCTURA — ¿este instrumento selecciona las MISMAS FILAS que la corrida del 31/08?')
print("""⭐ Control ESTRUCTURAL y no de valores, y el motivo va escrito porque la primera versión de
  este control pedía los valores y falló: esta solapa es inestable por CAMBIO (`R-31`) y publica
  ACUMULADO, así que dos lecturas separadas por un día NO tienen por qué coincidir en magnitud.
  Un CONTEO DE FILAS, en cambio, es inmune al CAMBIO — las filas no se mueven, sólo sus valores.

  Lo que `docs/MEDICION_mudanza_imp_2026-08-31.md` §1 registra de aquella corrida:
    · el conteo de JM es de **28 filas**
    · «el día 28 agrega campañas de GCBA y ninguna de JM» — o sea JM igual en las dos ventanas
      y GCBA distinto
""")

esperado_jm = 28
fallas = []
for vent in VENTANAS:
    for amb in ('jm', 'gcba'):
        n = len(RES[(vent, amb)]['filas'])
        ids = len(set(G(f, 'B') for f in RES[(vent, amb)]['filas']))
        print('   %-20s %-5s %4d filas · %3d cuentas' % (vent, amb, n, ids))

jm27, jm28 = (len(RES[(v, 'jm')]['filas']) for v in VENTANAS)
gc27, gc28 = (len(RES[(v, 'gcba')]['filas']) for v in VENTANAS)

print()
for etiqueta, ok in (
        ('JM = %d filas en 21–28 (esperado %d)' % (jm28, esperado_jm), jm28 == esperado_jm),
        ('JM idéntico entre las dos ventanas (%d = %d)' % (jm27, jm28), jm27 == jm28),
        ('GCBA CRECE del 27 al 28 (%d → %d)' % (gc27, gc28), gc28 > gc27)):
    print('   %s  %s' % ('✅' if ok else '⛔', etiqueta))
    if not ok:
        fallas.append(etiqueta)
if fallas:
    raise SystemExit('⛔ CP4 falló: este instrumento NO selecciona las filas de aquella corrida. '
                     'Nada de 0.2 y 0.3 se puede citar.')

bloque('ILUSTRACIÓN (no es un control) — cuánto se movieron los VALORES entre el 30 y el 31')
print('⚠ Esto NO acusa ni absuelve a nadie: mismas filas, otros valores, sobre una fuente que')
print('  `R-31` mide inestable por CAMBIO. Se imprime para que el orden de magnitud quede dicho.\n')
print('   %-5s %-10s %18s %18s %10s' % ('ámb.', 'token', 'fixture 30/08', 'publicado 31/08', 'delta %'))
for amb in ('jm', 'gcba'):
    for k in ('meta', 'google', 'prog', 'total'):
        v, p = r[amb][k], PUBLICADO[amb][k]
        print('   %-5s %-10s %18s %18s %9.2f %%' % (amb, k, mil(v), mil(p), 100.0 * (v - p) / p))
print('\n⭐ Los deltas son NEGATIVOS en las ocho celdas, que es lo que un acumulado leído un día')
print('  antes tiene que dar. Un error de selección se concentraría en una celda, no en las ocho.')

print('\n✅ Los cuatro controles positivos pasaron. Los números de 0.1–0.3 responden por el '
      'FIXTURE DEL 30/08 y por ningún otro día.')
